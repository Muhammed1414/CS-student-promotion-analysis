"""
Data utilities for student promotion analysis
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import plotly.express as px
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import train_test_split

def load_data(filepath):
    """Load student promotion dataset from CSV file"""
    return pd.read_csv(filepath)

def create_colorful_excel_report(df, output_path):
    """Create a colorful Excel report with conditional formatting"""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.formatting.rule import CellIsRule
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Promotion Data"
    
    # Add headers
    headers = list(df.columns)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add data
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Apply conditional formatting for Promotion Status
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    promotion_col = df.columns.get_loc("Promotion_Status") + 1
    last_row = len(df) + 1
    
    # Green for Promoted
    promoted_rule = CellIsRule(operator='equal', formula=['"Promoted"'], fill=green_fill)
    ws.conditional_formatting.add(f'O2:O{last_row}', promoted_rule)
    
    # Red for Not Promoted
    not_promoted_rule = CellIsRule(operator='equal', formula=['"Not Promoted"'], fill=red_fill)
    ws.conditional_formatting.add(f'O2:O{last_row}', not_promoted_rule)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_path)
    print(f"Colorful Excel report saved to {output_path}")

def generate_summary_stats(df):
    """Generate summary statistics for the dataset"""
    summary = {
        'Total Students': len(df),
        'Promoted Students': len(df[df['Promotion_Status'] == 'Promoted']),
        'Not Promoted Students': len(df[df['Promotion_Status'] == 'Not Promoted']),
        'Average GPA': df['GPA'].mean(),
        'Average Attendance': df['Attendance_Percentage'].mean(),
        'Average Courses Passed': df['Courses_Passed'].mean(),
        'Promotion Rate (%)': (len(df[df['Promotion_Status'] == 'Promoted']) / len(df)) * 100
    }
    return summary

def create_visualizations(df, output_dir):
    """Create colorful visualizations and save them"""
    import matplotlib.pyplot as plt
    import seaborn as sns
    
    # Set style
    plt.style.use('seaborn-v0_8')
    sns.set_palette("husl")
    
    # Create figure with subplots
    fig, axes = plt.subplots(2, 3, figsize=(18, 12))
    fig.suptitle('Computer Science Student Promotion Analysis', fontsize=20, fontweight='bold')
    
    # 1. Promotion Status Distribution
    promotion_counts = df['Promotion_Status'].value_counts()
    colors = ['#2E8B57', '#DC143C']
    axes[0,0].pie(promotion_counts.values, labels=promotion_counts.index, autopct='%1.1f%%', 
                  colors=colors, startangle=90)
    axes[0,0].set_title('Promotion Status Distribution', fontsize=14, fontweight='bold')
    
    # 2. GPA Distribution by Promotion Status
    df.boxplot(column='GPA', by='Promotion_Status', ax=axes[0,1])
    axes[0,1].set_title('GPA Distribution by Promotion Status', fontsize=14, fontweight='bold')
    axes[0,1].set_xlabel('Promotion Status')
    axes[0,1].set_ylabel('GPA')
    
    # 3. Attendance vs GPA Scatter Plot
    colors_map = {'Promoted': '#2E8B57', 'Not Promoted': '#DC143C'}
    scatter = axes[0,2].scatter(df['Attendance_Percentage'], df['GPA'], 
                                c=df['Promotion_Status'].map(colors_map), alpha=0.7)
    axes[0,2].set_xlabel('Attendance Percentage (%)')
    axes[0,2].set_ylabel('GPA')
    axes[0,2].set_title('Attendance vs GPA by Promotion Status', fontsize=14, fontweight='bold')
    
    # 4. Courses Passed Distribution
    axes[1,0].hist([df[df['Promotion_Status']=='Promoted']['Courses_Passed'], 
                    df[df['Promotion_Status']=='Not Promoted']['Courses_Passed']], 
                   bins=range(8, 15), label=['Promoted', 'Not Promoted'], alpha=0.7, color=['#2E8B57', '#DC143C'])
    axes[1,0].set_xlabel('Courses Passed')
    axes[1,0].set_ylabel('Number of Students')
    axes[1,0].set_title('Courses Passed Distribution', fontsize=14, fontweight='bold')
    axes[1,0].legend()
    
    # 5. Programming Skills vs Communication Skills
    axes[1,1].scatter(df['Programming_Skills'], df['Communication_Skills'], 
                      c=df['Promotion_Status'].map(colors_map), alpha=0.7)
    axes[1,1].set_xlabel('Programming Skills (1-10)')
    axes[1,1].set_ylabel('Communication Skills (1-10)')
    axes[1,1].set_title('Skills Comparison by Promotion Status', fontsize=14, fontweight='bold')
    
    # 6. Projects Completed Distribution
    axes[1,2].bar(['Promoted', 'Not Promoted'], 
                  [df[df['Promotion_Status']=='Promoted']['Projects_Completed'].mean(), 
                   df[df['Promotion_Status']=='Not Promoted']['Projects_Completed'].mean()],
                  color=['#2E8B57', '#DC143C'])
    axes[1,2].set_ylabel('Average Projects Completed')
    axes[1,2].set_title('Average Projects by Promotion Status', fontsize=14, fontweight='bold')
    
    plt.tight_layout()
    plt.savefig(f'{output_dir}/student_analysis_charts.png', dpi=300, bbox_inches='tight')
    plt.close()
    
    # Create correlation heatmap
    plt.figure(figsize=(10, 8))
    numeric_cols = ['GPA', 'Attendance_Percentage', 'Courses_Passed', 'Programming_Skills', 
                    'Communication_Skills', 'Projects_Completed']
    correlation_matrix = df[numeric_cols].corr()
    sns.heatmap(correlation_matrix, annot=True, cmap='RdYlBu_r', center=0)
    plt.title('Feature Correlation Heatmap', fontsize=16, fontweight='bold')
    plt.tight_layout()
    plt.savefig(f'{output_dir}/correlation_heatmap.png', dpi=300, bbox_inches='tight')
    plt.close()
    
    print(f"Visualizations saved to {output_dir}")
